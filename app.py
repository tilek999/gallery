from flask import Flask, render_template, request
from openpyxl import load_workbook


app=Flask(__name__)

@app.route("/")
def gallery():
    l = []
    excel = load_workbook("gallery.xlsx")
    page = excel["Sheet1"]
    for row in page:
        g = row[0].value
        c = row[1].value
        title = row[2].value
        lst = [title, g, c]
        l.append(lst)

    return render_template("gallery.html" ,l=l)

@app.route("/add-picture", methods=["POST"])
def add_picture():
    g = request.form.get("g")
    c = request.form.get("comment")
    title = request.form.get("title")
    
    excel = load_workbook("gallery.xlsx")
    sheet = excel["sheet1"]
    sheet.append(title, g, c)
    excel.save("gallery.xlsx")
    return render_template("gallery.html")
@app.route("/ad")
def add():
    return render_template("ad.html")    